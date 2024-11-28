# Обработка данных
import pandas as pd
import numpy as np

# Интерфейс и файловые диалоги
import tkinter as tk
from tkinter import filedialog

# Машинное обучение
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsRegressor
from sklearn.inspection import permutation_importance

# Работа с Excel
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook

# Сохранение и загрузка моделей
import joblib

# Работа с файловой системой
import os
from shutil import copyfile

# Визуализация
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

#Остальное
import math
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import mean_absolute_error

scaler = StandardScaler()



def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Расстояние между двумя точками на сфере по формуле Haversine.
    """
    R = 6371  # Радиус Земли в километрах
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

    dlat = lat2 - lat1
    dlon = lon2 - lon1

    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    distance = R * c * 1000  # Расстояние в метрах
    return round(distance, 1)  # Округляем до одного знака после запятой


# Загрузить данные и обучить LabelEncoder
data_study = pd.read_excel('C:\\Users\\andrew\\Downloads\\office_nn.xlsx')

# Создаём и обучаем LabelEncoder для столбца 'Функциональная зона'
label_encoder_1 = LabelEncoder()
label_encoder_1.fit(data_study['Функциональная зона'])

# Создаём и обучаем LabelEncoder для столбца 'Класс'
label_encoder_2 = LabelEncoder()
label_encoder_2.fit(data_study['Класс'])

def preprocessing_of_data(file_path):
    data = pd.read_excel(file_path)
    data = data.drop_duplicates()
    base_columns = [
    'Город', 'Адрес', 'Широта', 'Долгота', 'Функциональная зона', 
    'Общая площадь,\nкв.м', 'Год постройки', 'Материал стен', 'Класс', 
    'Тип здания', 'Этаж', 'Этажность', 'Состояние ремонта', 
    'Высота потолков, м', 'Планировка', 'Отдельный вход', 'Доступ', 
    'Охрана', 'Парковка', 'Красная линия']
    extra_columns = ['Цена предложения,\n руб.', 'Удельная цена, руб./кв.м']

    # Проверяем, присутствуют ли дополнительные столбцы в данных
    columns_to_check = base_columns + [col for col in extra_columns if col in data.columns]

    data = data.drop_duplicates(subset=columns_to_check, keep='first')

    columns_to_process = 'Цена предложения,\n руб.', 'Удельная цена, руб./кв.м'
    for column in columns_to_process:
        if column in data.columns:
            # Преобразуем в числовые данные
            data[column] = pd.to_numeric(data[column], errors='coerce')

    #преобразуем в булевый формат
    data['Охрана'] = data['Охрана'].replace({'есть': True, 'ЛОЖЬ': False})
    data['Охрана'] = pd.to_numeric(data['Охрана'], errors='coerce')
    data['Охрана'] = data['Охрана'].fillna(False) #заполняем пустые значения ложью
    data['Охрана'] = data['Охрана'].astype(bool)

    data['Отдельный вход'] = data['Отдельный вход'].replace({'Есть': True, 'нет': False})
    data['Отдельный вход'] = pd.to_numeric(data['Отдельный вход'], errors='coerce')
    data['Отдельный вход'] = data['Отдельный вход'].fillna(False)
    data['Отдельный вход'] = data['Отдельный вход'].astype(bool)

    data['Парковка'] = data['Парковка'].replace({'Есть': True, 'Организованная наземная открытая': True, 
                                                 'Стихийная наземная': True, 'Организованная подземная': True,
                                                 'Для грузового транспорта': True, 'Организованная наземная крытая': True })
    data['Парковка'] = pd.to_numeric(data['Парковка'], errors='coerce')
    data['Парковка'] = data['Парковка'].fillna(False)
    data['Парковка'] = data['Парковка'].astype(bool)

    data['Состояние ремонта'] = data['Состояние ремонта'].replace({'Типовой ремонт': True, 'Комфортный ремонт': True, 'Требуется косметический ремонт': False})
    data['Состояние ремонта'] = pd.to_numeric(data['Состояние ремонта'], errors='coerce')
    data['Состояние ремонта'] = data['Состояние ремонта'].fillna(False)
    data['Состояние ремонта'] = data['Состояние ремонта'].astype(bool)

    key_columns = ['Класс', 'Функциональная зона','Охрана', 'Парковка', 'Состояние ремонта', 'Отдельный вход', 'Широта', 'Долгота', 'Общая площадь,\nкв.м']
    data = data.dropna(subset=key_columns)
    
    data['Широта'] = pd.to_numeric(data['Широта'], errors='coerce')
    data['Долгота'] = pd.to_numeric(data['Долгота'], errors='coerce')
    data['Общая площадь,\nкв.м'] = pd.to_numeric(data['Общая площадь,\nкв.м'], errors='coerce')

    #  Применяем кодирование к столбцам в data_study
    data['Функциональная зона'] = label_encoder_1.transform(data['Функциональная зона'])
    data['Класс'] = label_encoder_2.transform(data['Класс'])

    return data

def find_best_k(X_train, X_test, Y_train, Y_test):
    # Параметры для сеточного поиска
    param_grid = {
        'n_neighbors': range(2, 52),  # Количество соседей
        'weights': ['uniform', 'distance'],  # Веса соседей
        'metric': ['euclidean', 'manhattan', 'chebyshev', 'minkowski'],  # Метрики расстояния
        'p': [1, 2],  # Параметр p для метрики minkowski (1 - манхэттенское, 2 - евклидово расстояние)
    }

    # Инициализация модели KNN
    knn = KNeighborsRegressor()

    # Инициализация GridSearchCV с использованием кросс-валидации
    grid_search = GridSearchCV(estimator=knn, param_grid=param_grid, cv=5, n_jobs=-1, scoring='neg_mean_absolute_error')

    # Обучение модели с сеточным поиском
    grid_search.fit(X_train, Y_train)

    # Лучшая комбинация гиперпараметров
    best_params = grid_search.best_params_
    best_k = best_params['n_neighbors']
    best_weights = best_params['weights']
    best_metric = best_params['metric']
    best_p = best_params['p']
    
    # Вывод результатов
    #print(f"Лучший k: {best_k}")
    #print(f"Лучшие веса: {best_weights}")
    #print(f"Лучшая метрика: {best_metric}")
    #print(f"Лучший p: {best_p}")

    # Прогнозирование с лучшими гиперпараметрами
    knn_best = KNeighborsRegressor(n_neighbors=best_k, weights=best_weights, metric=best_metric, p=best_p)
    knn_best.fit(X_train, Y_train)
    Y_pred = np.exp(knn_best.predict(X_test))

    # Оценка ошибки
    error = np.median(np.abs(np.exp(Y_test) - Y_pred) / np.exp(Y_test) * 100)
    print(f"Ошибка: {error}%")
    
    return best_k, best_weights, best_metric, best_p, error

def open_file_dialog(scaler):
    file_path = filedialog.askopenfilename(
        title="Выберите файл", 
        filetypes=(("Excel файлы", "*.xlsx *.xls"), ("Все файлы", "*.*"))
    )

    if file_path:
        file_path_converted = file_path.replace('/', '\\')  # Изменяем вид пути
        original_data = preprocessing_of_data('C:\\Users\\andrew\\Downloads\\office_nn.xlsx')
        data = preprocessing_of_data(file_path_converted)  # Используем выбранный файл

        # Проверяем наличие данных для расчета
        if not data['Удельная цена, руб./кв.м'].isnull().all() and (data['Удельная цена, руб./кв.м'] != 0).any():
            # Шаг 1: Выполняем предсказание
            predicted_values, distances, indices = knn_predict(data, original_data, scaler)
            
            # Шаг 2: Создаем таблицу и сохраняем в Excel
            result_table = create_excel_table(data, original_data, predicted_values, distances, indices)
            result_table_template = create_excel_table_template(result_table)
            # Выводим ошибку
            actual_values = data['Удельная цена, руб./кв.м'].values
            error_message = error_output(predicted_values, actual_values)
            text_output.delete(1.0, tk.END)  # Очищаем текстовое поле
            text_output.insert(tk.END, error_message)  # Вставляем текст ошибки
            
            # Строим график
            plot_feature_importance(data, scaler, frame)
        else: 
            # Создаем таблицу и сохраняем в Excel без заполнения столбца 'Изменение цены, %'
            predicted_values = pd.DataFrame({'predicted_values': [np.nan] * len(data)})
            result_table = create_excel_table(data, original_data, predicted_values, distances, indices)
            result_table_template = create_excel_table_template(result_table, calculate_change=False)
            text_output.delete(1.0, tk.END)  # Очищаем текстовое поле
            text_output.insert(tk.END, "Нет данных для вывода значения ошибки и построения графика приоритетов.")  # Сообщение об отсутствии данных


# Функция для изучения KNN
def knn_study(data):
    # Определяем признаки и целевую переменную
    X_target = data[['Класс', 'Функциональная зона', 'Охрана', 'Парковка', 'Состояние ремонта', 'Отдельный вход', 'Широта', 'Долгота', 'Общая площадь,\nкв.м']]
    Y_target = np.log(data['Удельная цена, руб./кв.м'])  # Используем одну квадратную скобку для выбора Series
    
    # Масштабируем данные
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_target)
    
    # Разделяем данные на обучающую и тестовую выборки
    X_learn, X_test, Y_learn, Y_test = train_test_split(X_scaled, Y_target, test_size=0.2, random_state=12)
    
    # Находим лучший K и ошибку, а также другие параметры
    best_k, best_weights, best_metric, best_p, error = find_best_k(X_learn, X_test, Y_learn, Y_test)
    
    # Создаем и обучаем KNN с найденными параметрами
    knn = KNeighborsRegressor(
        n_neighbors=best_k, 
        weights=best_weights, 
        metric=best_metric, 
        p=best_p
    )
    knn.fit(X_learn, Y_learn)
    
    print(f'Лучший k = {best_k}, Лучшие веса = {best_weights}, Лучшая метрика = {best_metric}, Лучший p = {best_p}, Ошибка = {error:.4f}%')
    
    # Сохраняем модель
    joblib.dump(knn, 'knn_model.joblib')
    
    return scaler

data_study = preprocessing_of_data('C:\\Users\\andrew\\Downloads\\office_nn.xlsx')
scaler = knn_study(data_study)

def knn_predict(data, original_data, scaler):
    """
    Функция для предсказания цен на основе KNN модели.
    Возвращает предсказания, исходные данные и индексы ближайших соседей.
    """
    # Загружаем модель
    knn = joblib.load('knn_model.joblib')
    
    # Определяем признаки
    X_target = data[['Класс', 'Функциональная зона', 'Охрана', 'Парковка', 'Состояние ремонта', 'Отдельный вход', 'Широта', 'Долгота', 'Общая площадь,\nкв.м']]
    # Масштабируем данные
    X_scaled = scaler.transform(X_target)
    
    # Получаем предсказания
    predicted_value = knn.predict(X_scaled)
    predicted_value = np.exp(predicted_value)  # Возвращаем обратно в исходные единицы

    # Вычисляем ошибку, если известны фактические цены
    if 'Удельная цена, руб./кв.м' in data.columns and not data['Удельная цена, руб./кв.м'].isnull().all():
        actual_prices = data['Удельная цена, руб./кв.м'].values
        error = np.median(np.abs(actual_prices - predicted_value) / actual_prices * 100)
        print(f'Ошибка предсказания: {error:.4f}%')

    # Получение расстояний и индексов ближайших соседей
    distances, indices = knn.kneighbors(X_scaled)
    
    return predicted_value, distances, indices

def create_excel_table(data, original_data, predicted_value, distances, indices, output_file='data.xlsx'):
    """
    Функция для создания Excel таблицы с предсказанными значениями и выделением строк.
    Сохраняет результат в Excel файл.
    """
    # Создание DataFrame для предсказанных значений
    predictions_df = pd.DataFrame(predicted_value, columns=['Предсказанная цена, руб/кв.м.'])
    new_rows = []
    k = len(indices[0])  # Количество соседей

    # Добавляем строки с предсказанными значениями
    for i in range(len(data)):
        row = data.iloc[i].copy()
        row['Предсказанная цена, руб/кв.м.'] = predictions_df.iloc[i, 0]
        new_rows.append(row)

        # Проверка на точное совпадение
        exact_match = original_data[
            (original_data[['Класс', 'Охрана', 'Парковка', 'Состояние ремонта', 'Отдельный вход', 'Широта', 'Долгота', 'Общая площадь,\nкв.м', 'Цена предложения,\n руб.', 'Удельная цена, руб./кв.м']] 
            == row[['Класс', 'Охрана', 'Парковка', 'Состояние ремонта', 'Отдельный вход', 'Широта', 'Долгота', 'Общая площадь,\nкв.м', 'Цена предложения,\n руб.', 'Удельная цена, руб./кв.м']].values).all(axis=1)
        ]

        if not exact_match.empty:
            match_row_copy = exact_match.iloc[0].copy()
            match_row_copy['Предсказанная цена, руб/кв.м.'] = None
            new_rows.append(match_row_copy)
            indices_neighbors = indices[i][1:k] 
        else:
            indices_neighbors = indices[i][:k]

         # Добавляем соседей и расстояния
        for j, neighbor_index in enumerate(indices_neighbors):
            neighbor_row = original_data.iloc[neighbor_index].copy()
            neighbor_row['Предсказанная цена, руб/кв.м.'] = None

            # Расчет расстояния до соседа
            distance = haversine_distance(
                row['Широта'], row['Долгота'],  # Координаты текущей точки
                neighbor_row['Широта'], neighbor_row['Долгота']  # Координаты соседа
            )
            neighbor_row['Расстояние, м'] = distance
            neighbor_row['Расстояние до соседа'] = distances[i][j]  # Добавление расстояния до соседа из KNN

            # Удаляем широту и долготу, если больше не нужны
            neighbor_row.drop(['Широта', 'Долгота'], inplace=True)

            # Добавляем строку соседа в new_rows
            new_rows.append(neighbor_row)



    # Создаём DataFrame из новых строк
    result_df = pd.DataFrame(new_rows)
    result_df.reset_index(drop=True, inplace=True)

    # Сохранение результата в Excel файл
    result_df.to_excel(output_file, index=False, engine='openpyxl')

    # Открытие Excel для выделения предсказанных строк
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for i in range(len(result_df)):
        if pd.notna(result_df.iloc[i]['Предсказанная цена, руб/кв.м.']):
            for col in range(1, len(result_df.columns) + 1):
                ws.cell(row=i + 2, column=col).fill = yellow_fill

    # Сохранение и закрытие файла
    wb.save(output_file)
    wb.close()
    #os.startfile(output_file)

    return result_df


def create_excel_table_template(result_df, template_file='Таблица шаблон.xlsx', output_table='Итог.xlsx', calculate_change=True):
    # Удаляем содержимое или создаём новый файл Итог.xlsx
    if os.path.exists(output_table):
        # Создаём новую рабочую книгу, чтобы очистить файл
        wb = Workbook()
        wb.save(output_table)
        wb.close()

    # Копируем шаблон в очищенный Итог.xlsx
    copyfile(template_file, output_table)
    wb = load_workbook(output_table)
    ws = wb.active

    # Загружаем первые две строки из шаблона
    template_wb = load_workbook(template_file)
    template_ws = template_wb.active

    # Копируем первые две строки из шаблона
    for row in template_ws.iter_rows(min_row=1, max_row=2):  # Только строки 1 и 2
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)  # Копируем значения
            if cell.has_style:
                new_cell._style = cell._style  # Копируем стиль

    template_wb.close()

    # Создаем новый DataFrame на основе result_df
    new_dataframe = result_df[['N', 'Ссылка', 'Сегмент', 'Тип сделки', 
                               'Город', 'Адрес', 'Функциональная зона', 'Расстояние, м',
                               'Этаж', 'Класс', 'Состояние ремонта', 'Общая площадь,\nкв.м',  
                               'Год постройки', 'Этажность', 'Материал стен', 'Охрана', 'Парковка',
                               'Удельная цена, руб./кв.м', 'Цена предложения,\n руб.', 
                               'Предсказанная цена, руб/кв.м.', 'Дата парсинга', 'Срок жизни/возраст объявления', 'Расстояние до соседа']]
    
    # Применяем декодирование для столбца 'Функциональная зона'
    new_dataframe['Функциональная зона'] = label_encoder_1.inverse_transform((new_dataframe['Функциональная зона']).values)

    # Применяем декодирование для столбца 'Класс'
    new_dataframe['Класс'] = label_encoder_2.inverse_transform((new_dataframe['Класс']).values)


    # Преобразуем значения в столбцах
    new_dataframe['Состояние ремонта'] = new_dataframe['Состояние ремонта'].replace({True: 'Есть', False: 'Нет'})
    new_dataframe['Парковка'] = new_dataframe['Парковка'].replace({True: 'Есть', False: 'Нет'})
    new_dataframe['Охрана'] = new_dataframe['Охрана'].replace({True: 'Есть', False: 'Нет'})

    # Добавляем новые столбцы
    new_dataframe.insert(4, 'Тип рынка', 'Офис')
    new_dataframe.insert(11, 'Количество комнат', 1)
    # Вставляем столбец 'Предсказанная цена, руб.' в нужную позицию
    new_dataframe.insert(22, 'Предсказанная цена, руб.', 
                        new_dataframe['Предсказанная цена, руб/кв.м.'] * new_dataframe['Общая площадь,\nкв.м'])

    # Вставляем столбец 'Изменение цены, %' после 'Предсказанная цена, руб.'
    if calculate_change:
        new_dataframe.insert(23, 'Изменение цены, %', 
                            ((new_dataframe['Предсказанная цена, руб/кв.м.'] - new_dataframe['Удельная цена, руб./кв.м']) / 
                            new_dataframe['Удельная цена, руб./кв.м']) * 100)
    else:
        new_dataframe.insert(23, 'Изменение цены, %', '-')

    

    # Добавляем строки new_dataframe в Excel, начиная с третьей строки
    start_row = 3
    for r_idx, row in enumerate(dataframe_to_rows(new_dataframe, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Выделяем предсказанные строки желтым цветом
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for i in range(len(new_dataframe)):
        if pd.notna(new_dataframe.iloc[i]['Предсказанная цена, руб/кв.м.']):
            for col in range(1, len(new_dataframe.columns) + 1):
                ws.cell(row=start_row + i, column=col).fill = yellow_fill

    # Генерируем гистограмму и сохраняем изображение
    histogram_filename = "histogram.png"
    plot_histogram(new_dataframe['Изменение цены, %'], histogram_filename)

    # Вставляем гистограмму в ячейку AE4
    insert_image_to_excel(ws, histogram_filename, 'AE4')

    # Сохранение и закрытие файла
    wb.save(output_table)
    wb.close()

    # Открытие нового файла для просмотра
    os.startfile(output_table)



def plot_histogram(data, filename):
    """Строим гистограмму и сохраняем её как изображение"""
    plt.figure(figsize=(8, 5))
    data = data.dropna()  # Удаляем NaN значения
    plt.hist(data, bins=20, color='lightcoral', edgecolor='black', alpha=0.75)
    plt.title('Распределение изменения цен, %')
    plt.xlabel('Изменение цены, %')
    plt.ylabel('Частота')
    plt.grid(True, linestyle='--', alpha=0.6)
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()


def insert_image_to_excel(ws, image_path, cell):
    """Вставка изображения в Excel"""
    img = Image(image_path)
    ws.add_image(img, cell)


def plot_feature_importance(data, scaler, frame): 
    # Определяем признаки
    X_target = data[['Класс', 'Охрана', 'Парковка', 'Состояние ремонта', 'Отдельный вход', 'Широта', 'Долгота', 'Общая площадь,\nкв.м']]
    
    # Масштабируем данные (используем transform)
    X_scaled = scaler.transform(X_target)

    y = data['Цена предложения,\n руб.']
    model = joblib.load('knn_model.joblib')

    # Оценка важности признаков с помощью Permutation Importance
    results = permutation_importance(model, X_scaled, y, n_repeats=30, random_state=42)
    
    # Получение важности признаков
    feature_importance = results.importances_mean
    feature_names = X_target.columns.tolist()

    # Очистка предыдущего графика
    plt.clf() 
    # Создание графика
    plt.figure(figsize=(10, 6))
    plt.barh(feature_names, feature_importance)
    plt.xlabel("Важность признаков")
    plt.title("Важность признаков для модели KNN")
    
    # Отображение графика в Tkinter
    canvas = FigureCanvasTkAgg(plt.gcf(), master=frame)  # Используем текущую фигуру
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.grid(row=3, column=0, padx=10, pady=10)  # Позиционирование графика в фрейме
    canvas.draw()  # Отображаем график

def error_output(predicted_values, actual_prices):
     error = np.median(np.abs(actual_prices - predicted_values) / actual_prices * 100)  
     error_message = f'Ошибка предсказания: {error:.4f}%'
     return error_message

# Создаём новый объект Tk
root = tk.Tk()
root.title("Knn Нижний Новгород")

# Настраиваем позицию на экране и размеры окна
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_widht = 800  # Увеличен размер окна для графика
window_height = 600
x_centre = (screen_width//2) - (window_widht//2)
y_centre = (screen_height//2) - (window_height//2)
root.geometry(f"{window_widht}x{window_height}+{x_centre}+{y_centre}")

# Создаём новый frame
frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

# Текстовая метка и кнопка выбора файла
label1 = tk.Label(frame, text="Это программа оценивает стоймость офисов. Выберите файл с офисами, которые нужно оценить", wraplength=300)
button1 = tk.Button(frame, text="Выберите файл excel", command=lambda: open_file_dialog(scaler))
label1.grid(row=0, column=0, padx=10)
button1.grid(row=1, column=0, padx=10, pady=15)

# Поле для вывода текста
text_output = tk.Text(frame, height=10, width=50)
text_output.grid(row=2, column=0, padx=10, pady=10)


root.mainloop()

